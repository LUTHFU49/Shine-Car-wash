"""
Generic tabular export helpers shared by every report view. Each report
in views.py builds a plain `headers` list and `rows` list (list of
lists, already display-formatted) and hands them to one of these --
the same styling conventions as apps.inventory.reports and
apps.payments.reports (same brand colors, same ReportLab table style),
just parameterized instead of duplicated ten times over.
"""
from apps.core.csv_utils import safe_csv_writer

from django.conf import settings
from django.http import HttpResponse

BRAND_BLUE = '#0013DE'


def csv_response(filename, headers, rows, summary_lines=None):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    writer = safe_csv_writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    if summary_lines:
        writer.writerow([])
        for line in summary_lines:
            writer.writerow([line])
    return response


def excel_response(filename, sheet_title, headers, rows, summary_lines=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title[:31]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append(row)

    if summary_lines:
        sheet.append([])
        for line in summary_lines:
            sheet.append([line])
            sheet.cell(row=sheet.max_row, column=1).font = Font(bold=True, italic=True)

    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = max(12, length + 2)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response


def _pdf_base_doc(response, title):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate

    return SimpleDocTemplate(
        response, pagesize=A4,
        topMargin=18 * mm, bottomMargin=18 * mm, leftMargin=16 * mm, rightMargin=16 * mm,
        title=title,
    )


def _pdf_header_elements(title, subtitle=''):
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer

    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'<b>{getattr(settings, "SITE_NAME", "ShineHub")}</b>', styles['Title']),
        Paragraph(getattr(settings, 'COMPANY_NAME', ''), styles['Normal']),
        Spacer(1, 10),
        Paragraph(f'<b>{title}</b>', styles['Heading2']),
    ]
    if subtitle:
        elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 12))
    return elements


def _styled_table(data, col_widths=None):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(BRAND_BLUE)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8.5),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E4E7F1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F6FB')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    return table


def pdf_response(filename, title, subtitle, headers, rows, summary_lines=None):
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

    doc = _pdf_base_doc(response, title)
    elements = _pdf_header_elements(title, subtitle)
    elements.append(_styled_table([headers, *rows]))

    if summary_lines:
        styles = getSampleStyleSheet()
        elements.append(Spacer(1, 14))
        for line in summary_lines:
            elements.append(Paragraph(f'<b>{line}</b>', styles['Normal']))

    doc.build(elements)
    return response
