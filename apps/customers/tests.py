from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import Role
from apps.audit_logs.models import AuditLog

from .models import Customer

User = get_user_model()


def make_staff_user(role=Role.MANAGER, username='staffuser'):
    user = User.objects.create_user(
        username=username, email=f'{username}@example.com', password='StrongPass1!', role=role,
    )
    user.is_email_verified = True
    user.save()
    return user


def make_customer_user(username='customeruser'):
    return User.objects.create_user(
        username=username, email=f'{username}@example.com', password='StrongPass1!', role=Role.CUSTOMER,
        first_name='Auto', last_name='Created', phone_number='0711111111',
    )


class CustomerModelTests(TestCase):
    def test_customer_code_format(self):
        customer = Customer.objects.create(first_name='Jane', last_name='Doe', phone_number='0722222222')
        self.assertTrue(customer.customer_code.startswith('CUST-'))
        self.assertEqual(customer.customer_code, f'CUST-{customer.pk:06d}')

    def test_full_name_property(self):
        customer = Customer.objects.create(first_name='Jane', last_name='Doe', phone_number='0722222223')
        self.assertEqual(customer.full_name, 'Jane Doe')

    def test_is_linked_account_false_for_walkin(self):
        customer = Customer.objects.create(first_name='Jane', last_name='Doe', phone_number='0722222224')
        self.assertFalse(customer.is_linked_account)

    def test_phone_number_must_be_unique(self):
        Customer.objects.create(first_name='Jane', last_name='Doe', phone_number='0722222225')
        with self.assertRaises(Exception):
            Customer.objects.create(first_name='John', last_name='Smith', phone_number='0722222225')


class AutoCustomerProfileSignalTests(TestCase):
    def test_self_registration_creates_linked_customer(self):
        user = make_customer_user()
        customer = Customer.objects.get(user=user)
        self.assertEqual(customer.full_name, 'Auto Created')
        self.assertTrue(customer.is_linked_account)
        self.assertEqual(customer.phone_number, '0711111111')

    def test_non_customer_role_does_not_create_customer_profile(self):
        make_staff_user(role=Role.EMPLOYEE, username='emp1')
        self.assertEqual(Customer.objects.count(), 0)


class CustomerPermissionTests(TestCase):
    """Only Super Admin / Manager / Cashier may access the Customers app."""

    def setUp(self):
        self.customer = Customer.objects.create(first_name='Jane', last_name='Doe', phone_number='0733333331')
        self.list_url = reverse('customers:list')

    def test_anonymous_redirected_to_login(self):
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('accounts:login'), response.url)

    def test_customer_role_gets_403(self):
        make_customer_user(username='cust1')
        self.client.login(username='cust1', password='StrongPass1!')
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 403)

    def test_employee_role_gets_403(self):
        make_staff_user(role=Role.EMPLOYEE, username='emp2')
        self.client.login(username='emp2', password='StrongPass1!')
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 403)

    def test_manager_role_allowed(self):
        make_staff_user(role=Role.MANAGER, username='mgr1')
        self.client.login(username='mgr1', password='StrongPass1!')
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 200)

    def test_cashier_role_allowed(self):
        make_staff_user(role=Role.CASHIER, username='cash1')
        self.client.login(username='cash1', password='StrongPass1!')
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 200)

    def test_super_admin_role_allowed(self):
        make_staff_user(role=Role.SUPER_ADMIN, username='admin1')
        self.client.login(username='admin1', password='StrongPass1!')
        response = self.client.get(self.list_url)
        self.assertEqual(response.status_code, 200)


class CustomerCRUDTests(TestCase):
    def setUp(self):
        self.staff = make_staff_user()
        self.client.login(username='staffuser', password='StrongPass1!')

    def test_create_customer(self):
        response = self.client.post(reverse('customers:create'), {
            'first_name': 'Grace', 'last_name': 'Njeri', 'phone_number': '0744444441',
            'email': 'grace@example.com',
        }, follow=True)
        self.assertEqual(response.status_code, 200)
        customer = Customer.objects.get(phone_number='0744444441')
        self.assertEqual(customer.created_by, self.staff)
        self.assertTrue(
            AuditLog.objects.filter(model_name='Customer', object_id=str(customer.pk), action=AuditLog.Action.CREATE).exists()
        )

    def test_create_customer_rejects_duplicate_phone(self):
        Customer.objects.create(first_name='A', last_name='B', phone_number='0744444442')
        response = self.client.post(reverse('customers:create'), {
            'first_name': 'C', 'last_name': 'D', 'phone_number': '0744444442',
        })
        self.assertEqual(response.status_code, 200)  # re-rendered with errors
        self.assertEqual(Customer.objects.filter(phone_number='0744444442').count(), 1)

    def test_create_customer_rejects_invalid_name(self):
        response = self.client.post(reverse('customers:create'), {
            'first_name': 'Grace123', 'last_name': 'Njeri', 'phone_number': '0744444443',
        })
        self.assertEqual(response.status_code, 200)
        self.assertFalse(Customer.objects.filter(phone_number='0744444443').exists())

    def test_edit_customer(self):
        customer = Customer.objects.create(first_name='Grace', last_name='Njeri', phone_number='0744444444')
        response = self.client.post(reverse('customers:edit', args=[customer.public_id]), {
            'first_name': 'Grace', 'last_name': 'Wanjiru', 'phone_number': '0744444444',
        }, follow=True)
        self.assertEqual(response.status_code, 200)
        customer.refresh_from_db()
        self.assertEqual(customer.last_name, 'Wanjiru')

    def test_detail_view_shows_history(self):
        customer = Customer.objects.create(first_name='Grace', last_name='Njeri', phone_number='0744444445', created_by=self.staff)
        AuditLog.objects.create(user=self.staff, action=AuditLog.Action.CREATE, model_name='Customer', object_id=str(customer.pk), description='Test event')
        response = self.client.get(reverse('customers:detail', args=[customer.public_id]))
        self.assertContains(response, 'Test event')

    def test_deactivate_and_reactivate(self):
        customer = Customer.objects.create(first_name='Grace', last_name='Njeri', phone_number='0744444446')
        self.client.post(reverse('customers:deactivate', args=[customer.public_id]))
        customer.refresh_from_db()
        self.assertFalse(customer.is_active)
        self.assertIsNotNone(customer.deactivated_at)

        self.client.post(reverse('customers:reactivate', args=[customer.public_id]))
        customer.refresh_from_db()
        self.assertTrue(customer.is_active)
        self.assertIsNone(customer.deactivated_at)

    def test_deactivate_requires_post(self):
        customer = Customer.objects.create(first_name='Grace', last_name='Njeri', phone_number='0744444447')
        response = self.client.get(reverse('customers:deactivate', args=[customer.public_id]))
        self.assertEqual(response.status_code, 405)


class CustomerSearchAndFilterTests(TestCase):
    def setUp(self):
        make_staff_user()
        self.client.login(username='staffuser', password='StrongPass1!')
        self.active = Customer.objects.create(first_name='Alice', last_name='Wambui', phone_number='0755555551', email='alice@example.com')
        self.inactive = Customer.objects.create(first_name='Bob', last_name='Otieno', phone_number='0755555552', is_active=False)

    def test_search_by_name(self):
        response = self.client.get(reverse('customers:list'), {'q': 'Alice'})
        self.assertContains(response, 'Alice Wambui')
        self.assertNotContains(response, 'Bob Otieno')

    def test_search_by_phone(self):
        response = self.client.get(reverse('customers:list'), {'q': '0755555552'})
        self.assertContains(response, 'Bob Otieno')

    def test_search_by_customer_code(self):
        code = self.active.customer_code
        response = self.client.get(reverse('customers:list'), {'q': code})
        self.assertContains(response, 'Alice Wambui')

    def test_filter_by_status_active(self):
        response = self.client.get(reverse('customers:list'), {'status': 'active'})
        self.assertContains(response, 'Alice Wambui')
        self.assertNotContains(response, 'Bob Otieno')

    def test_filter_by_status_inactive(self):
        response = self.client.get(reverse('customers:list'), {'status': 'inactive'})
        self.assertContains(response, 'Bob Otieno')
        self.assertNotContains(response, 'Alice Wambui')

    def test_filter_by_source_linked(self):
        make_customer_user(username='linkedcust')
        response = self.client.get(reverse('customers:list'), {'source': 'linked'})
        self.assertContains(response, 'Auto Created')
        self.assertNotContains(response, 'Alice Wambui')


class CustomerExportTests(TestCase):
    def setUp(self):
        make_staff_user()
        self.client.login(username='staffuser', password='StrongPass1!')
        Customer.objects.create(first_name='Export', last_name='Test', phone_number='0766666661', email='export@example.com')

    def test_csv_export(self):
        response = self.client.get(reverse('customers:export_csv'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'text/csv')
        content = response.content.decode()
        self.assertIn('Export', content)
        self.assertIn('Test', content)
        self.assertIn('0766666661', content)
        self.assertIn('export@example.com', content)

    def test_excel_export(self):
        response = self.client.get(reverse('customers:export_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertTrue(len(response.content) > 0)


class CustomerExportCSVInjectionTests(TestCase):
    """Phase 15: confirm the safe_csv_writer swap is actually wired up here."""

    def test_malicious_customer_name_is_neutralized_in_export(self):
        import csv
        import io

        staff = make_staff_user()
        self.client.login(username=staff.username, password='StrongPass1!')
        Customer.objects.create(first_name='=cmd|test', last_name='Doe', email='x@example.com', phone_number='0700000000')

        response = self.client.get(reverse('customers:export_csv'))
        rows = list(csv.reader(io.StringIO(response.content.decode())))
        first_name_cell = rows[1][1]
        self.assertTrue(first_name_cell.startswith("'="))

    def test_malicious_customer_name_is_neutralized_in_excel_export(self):
        from openpyxl import load_workbook
        import io

        staff = make_staff_user()
        self.client.login(username=staff.username, password='StrongPass1!')
        Customer.objects.create(first_name='=cmd|test', last_name='Doe', email='y@example.com', phone_number='0700000001')

        response = self.client.get(reverse('customers:export_excel'))
        wb = load_workbook(io.BytesIO(response.content))
        sheet = wb.active
        first_name_cell = sheet.cell(row=2, column=2).value
        self.assertTrue(str(first_name_cell).startswith("'="))
